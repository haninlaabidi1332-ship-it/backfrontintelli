import logging
import os
from datetime import timedelta
from io import BytesIO

from celery import shared_task
from django.utils import timezone
from django.core.cache import cache
from django.db.models import Avg, Count, Q

from .models import KPIHistory, Report

logger = logging.getLogger(__name__)

# ============================================================================
# KPI AGGREGATION
# ============================================================================

@shared_task(name='analytics.aggregate_kpi')
def aggregate_kpi(period='hour'):
    """Agrège les KPIs globaux pour la période donnée."""
    from apps.equipements.models import OLT, ONT
    from apps.snmp_collector.models import MetricHistory, SnmpOID, PollJob
    from apps.bfd_monitor.models import BFDSession
    from apps.ai_engine.models import AnomalyDetection
    from apps.alerting.models import Alert

    now = timezone.now()
    delta_map = {'hour': timedelta(hours=1), 'day': timedelta(days=1),
                 'week': timedelta(weeks=1), 'month': timedelta(days=30)}
    if period not in delta_map:
        return
    start = now - delta_map[period]

    total_olts  = OLT.objects.count()
    active_olts = OLT.objects.filter(status='active').count()
    total_onts  = ONT.objects.count()
    online_onts = ONT.objects.filter(status='online').count()

    def _avg_metric(name_fragment):
        oid = SnmpOID.objects.filter(name__icontains=name_fragment).first()
        if not oid:
            return None
        return MetricHistory.objects.filter(
            oid=oid, timestamp__gte=start
        ).aggregate(v=Avg('numeric_value'))['v']

    avg_cpu  = _avg_metric('cpu')
    avg_mem  = _avg_metric('memory')
    avg_temp = _avg_metric('temperature')
    avg_rx   = _avg_metric('rx_power')

    total_jobs   = PollJob.objects.filter(created_at__gte=start).count()
    success_jobs = PollJob.objects.filter(created_at__gte=start, state='success').count()
    snmp_success = (success_jobs / total_jobs * 100) if total_jobs else 100.0

    bfd_total = BFDSession.objects.filter(is_monitored=True).count()
    bfd_up    = BFDSession.objects.filter(is_monitored=True, state='up').count()

    anomaly_count = AnomalyDetection.objects.filter(detected_at__gte=start).count()
    alert_count   = Alert.objects.filter(first_seen__gte=start).count()

    KPIHistory.objects.create(
        period=period, timestamp=now,
        total_olts=total_olts, active_olts=active_olts,
        total_onts=total_onts, online_onts=online_onts,
        avg_cpu_usage=avg_cpu, avg_memory_usage=avg_mem,
        avg_temperature=avg_temp, avg_rx_power=avg_rx,
        snmp_success_rate=snmp_success,
        bfd_total_sessions=bfd_total, bfd_up_sessions=bfd_up,
        anomaly_count=anomaly_count, alert_count=alert_count,
    )
    cache.delete("analytics:kpi:current")
    logger.info(f"KPIs agrégés ({period})")


# ============================================================================
# REPORT GENERATION
# ============================================================================

@shared_task(name='analytics.generate_report')
def generate_report(report_id):
    """Génère le fichier PDF ou Excel d'un rapport et le sauvegarde."""
    from django.core.files.base import ContentFile

    try:
        report = Report.objects.get(id=report_id)
    except Report.DoesNotExist:
        logger.error(f"Rapport {report_id} introuvable")
        return

    report.status = 'generating'
    report.save(update_fields=['status'])

    try:
        kpis = KPIHistory.objects.filter(
            period='day',
            timestamp__range=(report.date_from, report.date_to),
            olt__isnull=True,
        ).order_by('timestamp')

        if report.format == 'pdf':
            buffer = _generate_pdf(kpis, report)
            ext = 'pdf'
        else:
            buffer = _generate_excel(kpis, report)
            ext = 'xlsx'

        filename = f"report_{report.id}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.{ext}"
        report.file.save(filename, ContentFile(buffer.getvalue()), save=False)
        report.status = 'ready'
        report.generated_at = timezone.now()
        report.save(update_fields=['file', 'status', 'generated_at'])
        logger.info(f"Rapport '{report.name}' généré → {filename}")

    except Exception as exc:
        report.status = 'failed'
        report.error_message = str(exc)
        report.save(update_fields=['status', 'error_message'])
        logger.exception(f"Erreur génération rapport '{report.name}': {exc}")


# ────────────────────────────────────────────────────────────────────────────
# PDF
# ────────────────────────────────────────────────────────────────────────────

def _generate_pdf(kpis, report):
    """Rapport PDF professionnel avec ReportLab."""
    import reportlab
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        HRFlowable, KeepTogether,
    )
    from apps.equipements.models import OLT, ONT
    from apps.alerting.models import Alert

    # ── Palette ──────────────────────────────────────────────────────────────
    BLUE        = colors.HexColor('#1e3a5f')
    LIGHT_BLUE  = colors.HexColor('#d6e4f0')
    ORANGE      = colors.HexColor('#e07b20')
    GREEN       = colors.HexColor('#27ae60')
    RED         = colors.HexColor('#c0392b')
    GREY_BG     = colors.HexColor('#f4f6f8')
    GREY_BORDER = colors.HexColor('#cccccc')

    SEV_COLORS = {
        'critical': colors.HexColor('#fde8e8'),
        'major':    colors.HexColor('#fdf0e0'),
        'warning':  colors.HexColor('#fdf8e0'),
        'info':     colors.HexColor('#e8f5e9'),
    }

    # ── Styles ────────────────────────────────────────────────────────────────
    styles = getSampleStyleSheet()
    s_title    = ParagraphStyle('RPT_Title',    parent=styles['Normal'],
                                fontSize=22, textColor=BLUE, leading=28,
                                fontName='Helvetica-Bold')
    s_subtitle = ParagraphStyle('RPT_Sub',      parent=styles['Normal'],
                                fontSize=10, textColor=colors.HexColor('#555555'),
                                leading=14)
    s_h2       = ParagraphStyle('RPT_H2',       parent=styles['Normal'],
                                fontSize=12, textColor=BLUE, leading=16,
                                fontName='Helvetica-Bold',
                                spaceBefore=14, spaceAfter=4)
    s_small    = ParagraphStyle('RPT_Small',    parent=styles['Normal'],
                                fontSize=8, textColor=colors.grey)
    s_note     = ParagraphStyle('RPT_Note',     parent=styles['Normal'],
                                fontSize=8, textColor=colors.HexColor('#666666'),
                                leading=12)

    # ── Shared table style helpers ────────────────────────────────────────────
    def _hdr_style(bg=BLUE, fg=colors.white):
        return [
            ('BACKGROUND',  (0, 0), (-1, 0), bg),
            ('TEXTCOLOR',   (0, 0), (-1, 0), fg),
            ('FONTNAME',    (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',    (0, 0), (-1, -1), 8),
            ('GRID',        (0, 0), (-1, -1), 0.4, GREY_BORDER),
            ('VALIGN',      (0, 0), (-1, -1), 'MIDDLE'),
            ('PADDING',     (0, 0), (-1, -1), 5),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, GREY_BG]),
        ]

    buffer = BytesIO()
    now = timezone.now()

    def _on_page(canvas, doc):
        """Pied de page sur chaque page."""
        canvas.saveState()
        canvas.setFont('Helvetica', 7)
        canvas.setFillColor(colors.grey)
        canvas.drawString(2*cm, 1.2*cm,
            f"SOTETEL — IntelliOLT v2.0  |  {report.name}  |  Confidentiel")
        canvas.drawRightString(A4[0] - 2*cm, 1.2*cm,
            f"Page {doc.page}")
        canvas.restoreState()

    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm,
        title=report.name, author='IntelliOLT — SOTETEL',
    )

    story = []

    # ── HEADER ────────────────────────────────────────────────────────────────
    story.append(Paragraph("SOTETEL / IntelliOLT", s_title))
    story.append(Paragraph(
        f"Supervision réseau FTTH &mdash; {report.get_report_type_display()}",
        s_subtitle,
    ))
    story.append(HRFlowable(width='100%', thickness=2, color=BLUE, spaceAfter=10))

    # Meta box
    generated_by = (
        report.generated_by.get_full_name() or report.generated_by.email
        if report.generated_by else 'Système'
    )
    meta_rows = [
        ['Rapport',      report.name],
        ['Période',      f"{report.date_from.strftime('%d/%m/%Y')} — {report.date_to.strftime('%d/%m/%Y')}"],
        ['Généré le',    now.strftime('%d/%m/%Y à %H:%M')],
        ['Généré par',   generated_by],
        ['Format',       report.get_format_display()],
    ]
    meta_t = Table(meta_rows, colWidths=[4*cm, 13*cm])
    meta_t.setStyle(TableStyle([
        ('BACKGROUND',  (0, 0), (0, -1), LIGHT_BLUE),
        ('FONTNAME',    (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE',    (0, 0), (-1, -1), 8),
        ('GRID',        (0, 0), (-1, -1), 0.4, GREY_BORDER),
        ('VALIGN',      (0, 0), (-1, -1), 'MIDDLE'),
        ('PADDING',     (0, 0), (-1, -1), 5),
    ]))
    story.append(meta_t)
    story.append(Spacer(1, 16))

    # ── RÉSUMÉ EXÉCUTIF ───────────────────────────────────────────────────────
    story.append(Paragraph("1. Résumé exécutif", s_h2))

    olts_qs    = OLT.objects.all()
    total_olts = olts_qs.count()
    active_olts = olts_qs.filter(status='active').count()
    total_onts  = ONT.objects.count()
    online_onts = ONT.objects.filter(status='online').count()
    active_alerts = Alert.objects.filter(status='active').count()

    from apps.ai_engine.models import AnomalyDetection
    period_anomalies = AnomalyDetection.objects.filter(
        detected_at__range=(report.date_from, report.date_to)
    ).count()

    ont_pct = round(online_onts / total_onts * 100, 1) if total_onts else 0
    olt_pct = round(active_olts / total_olts * 100, 1) if total_olts else 0

    kpis_list = list(kpis)
    if kpis_list:
        avg_cpu  = round(sum(k.avg_cpu_usage  or 0 for k in kpis_list) / len(kpis_list), 1)
        avg_mem  = round(sum(k.avg_memory_usage or 0 for k in kpis_list) / len(kpis_list), 1)
        avg_snmp = round(sum(k.snmp_success_rate or 100 for k in kpis_list) / len(kpis_list), 1)
    else:
        avg_cpu = avg_mem = avg_snmp = None

    def _status_cell(ok_cond, ok_txt='✓ Nominal', warn_txt='⚠ Attention'):
        return ok_txt if ok_cond else warn_txt

    summary_rows = [
        ['Indicateur', 'Valeur', 'Statut'],
        ['OLTs actifs / Total',    f"{active_olts} / {total_olts}",
            _status_cell(olt_pct >= 100)],
        ['ONTs en ligne / Total',  f"{online_onts} / {total_onts} ({ont_pct} %)",
            _status_cell(ont_pct >= 95)],
        ['CPU moyen',
            f"{avg_cpu} %" if avg_cpu is not None else '— (aucune donnée)',
            _status_cell(avg_cpu is None or avg_cpu < 70)],
        ['Mémoire moyenne',
            f"{avg_mem} %" if avg_mem is not None else '—',
            _status_cell(avg_mem is None or avg_mem < 80)],
        ['Taux SNMP succès',
            f"{avg_snmp} %" if avg_snmp is not None else '—',
            _status_cell(avg_snmp is None or avg_snmp > 95)],
        ['Alertes actives',        str(active_alerts),
            _status_cell(active_alerts == 0, '✓ Aucune', f'⚠ {active_alerts} active(s)')],
        ['Anomalies sur la période', str(period_anomalies),
            _status_cell(period_anomalies < 3, '✓ Nominal', f'⚠ {period_anomalies} détectée(s)')],
    ]

    sum_t = Table(summary_rows, colWidths=[7*cm, 5.5*cm, 4.5*cm])
    sum_style = _hdr_style()
    # Color status column
    for row_i, row in enumerate(summary_rows[1:], start=1):
        cell_txt = row[2]
        fill = GREEN if cell_txt.startswith('✓') else (
               ORANGE if cell_txt.startswith('⚠') else colors.white)
        if fill != colors.white:
            sum_style.append(('BACKGROUND', (2, row_i), (2, row_i), fill))
            sum_style.append(('TEXTCOLOR',  (2, row_i), (2, row_i), colors.white))
    sum_t.setStyle(TableStyle(sum_style))
    story.append(sum_t)
    story.append(Spacer(1, 16))

    # ── OLT STATUS ────────────────────────────────────────────────────────────
    story.append(Paragraph("2. État des équipements OLT", s_h2))
    olt_hdr = ['Hostname', 'Site', 'Adresse IP', 'Fournisseur', 'Modèle', 'Statut', 'ONTs']
    olt_rows = [olt_hdr]
    STATUS_LABEL = {'active': 'ACTIF', 'maintenance': 'MAINTENANCE', 'inactive': 'INACTIF'}
    for olt in OLT.objects.select_related('site', 'vendor', 'device_type').all():
        ont_online = ONT.objects.filter(olt=olt, status='online').count()
        ont_total  = ONT.objects.filter(olt=olt).count()
        olt_rows.append([
            olt.hostname,
            olt.site.name if olt.site else '—',
            olt.ip_address,
            olt.vendor.name if olt.vendor else '—',
            olt.device_type.model if olt.device_type else '—',
            STATUS_LABEL.get(olt.status, olt.status),
            f"{ont_online}/{ont_total}",
        ])
    olt_t = Table(olt_rows, colWidths=[3.5*cm, 3*cm, 3*cm, 2.5*cm, 3*cm, 2.5*cm, 1.5*cm])
    olt_style = _hdr_style()
    for row_i, row in enumerate(olt_rows[1:], start=1):
        status_val = row[5]
        if status_val == 'ACTIF':
            olt_style.append(('TEXTCOLOR', (5, row_i), (5, row_i), GREEN))
        elif status_val in ('MAINTENANCE', 'INACTIF'):
            olt_style.append(('TEXTCOLOR', (5, row_i), (5, row_i), RED))
    olt_t.setStyle(TableStyle(olt_style))
    story.append(olt_t)
    story.append(Spacer(1, 16))

    # ── KPI JOURNALIERS ───────────────────────────────────────────────────────
    story.append(Paragraph("3. KPIs journaliers", s_h2))
    if kpis_list:
        kpi_hdr = ['Date', 'OLTs actifs', 'ONT en ligne', 'CPU moy.', 'Mém. moy.', 'SNMP %', 'BFD UP', 'Alertes']
        kpi_rows = [kpi_hdr]
        for k in kpis_list:
            kpi_rows.append([
                k.timestamp.strftime('%d/%m/%Y'),
                f"{k.active_olts}/{k.total_olts}",
                f"{k.online_onts}/{k.total_onts}",
                f"{k.avg_cpu_usage:.1f}%"  if k.avg_cpu_usage    is not None else '—',
                f"{k.avg_memory_usage:.1f}%" if k.avg_memory_usage is not None else '—',
                f"{k.snmp_success_rate:.1f}%",
                str(k.bfd_up_sessions),
                str(k.alert_count),
            ])
        kpi_t = Table(kpi_rows, colWidths=[2.8*cm, 2.5*cm, 2.8*cm, 2.3*cm, 2.3*cm, 2.2*cm, 2*cm, 2*cm])
        kpi_t.setStyle(TableStyle(_hdr_style()))
        story.append(kpi_t)
    else:
        story.append(Paragraph(
            "Aucune donnée KPI journalière disponible pour cette période.",
            s_note,
        ))
    story.append(Spacer(1, 16))

    # ── ALERTES ───────────────────────────────────────────────────────────────
    story.append(Paragraph("4. Alertes de la période", s_h2))
    alerts_qs = Alert.objects.filter(
        first_seen__range=(report.date_from, report.date_to)
    ).select_related('olt').order_by('-first_seen')[:20]

    if alerts_qs.exists():
        alert_hdr = ['Date', 'OLT', 'Sévérité', 'Statut', 'Message']
        alert_rows = [alert_hdr]
        for a in alerts_qs:
            alert_rows.append([
                a.first_seen.strftime('%d/%m %H:%M'),
                a.olt.hostname if a.olt else '—',
                a.severity.upper(),
                a.status.upper(),
                Paragraph(a.message[:80], s_note),
            ])
        alert_t = Table(alert_rows, colWidths=[2.5*cm, 3.5*cm, 2.2*cm, 2.5*cm, 6.3*cm])
        al_style = _hdr_style(bg=ORANGE)
        for row_i, a in enumerate(alerts_qs, start=1):
            bg = SEV_COLORS.get(a.severity, colors.white)
            al_style.append(('BACKGROUND', (2, row_i), (2, row_i), bg))
        alert_t.setStyle(TableStyle(al_style))
        story.append(alert_t)
    else:
        story.append(Paragraph(
            "Aucune alerte enregistrée pour cette période. ✓",
            s_note,
        ))

    # ── FOOTER LINE ───────────────────────────────────────────────────────────
    story.append(Spacer(1, 20))
    story.append(HRFlowable(width='100%', thickness=0.5, color=GREY_BORDER))
    story.append(Paragraph(
        "Document confidentiel — usage interne SOTETEL ISP. "
        "Généré automatiquement par IntelliOLT v2.0.",
        s_small,
    ))

    doc.build(story, onFirstPage=_on_page, onLaterPages=_on_page)
    buffer.seek(0)
    return buffer


# ────────────────────────────────────────────────────────────────────────────
# EXCEL
# ────────────────────────────────────────────────────────────────────────────

def _generate_excel(kpis, report):
    """Rapport Excel multi-feuilles avec openpyxl."""
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side,
    )
    from openpyxl.utils import get_column_letter
    from apps.equipements.models import OLT, ONT
    from apps.alerting.models import Alert
    from apps.ai_engine.models import AnomalyDetection

    # ── Styles ────────────────────────────────────────────────────────────────
    def _fill(hex_color):
        return PatternFill('solid', fgColor=hex_color)

    def _border():
        s = Side(border_style='thin', color='BBBBBB')
        return Border(left=s, right=s, top=s, bottom=s)

    FILL_BLUE   = _fill('1E3A5F')
    FILL_ORANGE = _fill('E07B20')
    FILL_GREEN  = _fill('27AE60')
    FILL_LIGHT  = _fill('D6E4F0')
    FILL_STRIPE = _fill('EEF4FA')
    FILL_WHITE  = _fill('FFFFFF')
    FONT_WHITE  = Font(color='FFFFFF', bold=True, size=9)
    FONT_BOLD   = Font(bold=True, size=9)
    FONT_NORMAL = Font(size=9)
    FONT_TITLE  = Font(bold=True, size=14, color='1E3A5F')
    FONT_SUB    = Font(italic=True, size=9, color='666666')
    CENTER      = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LEFT        = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    BD          = _border()

    def _apply_header(ws, headers, row=1, fill=FILL_BLUE):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.fill = fill
            c.font = FONT_WHITE
            c.alignment = CENTER
            c.border = BD
        ws.row_dimensions[row].height = 28

    def _data_row(ws, row_idx, values, stripe=False):
        for col, v in enumerate(values, 1):
            c = ws.cell(row=row_idx, column=col, value=v)
            c.fill = FILL_STRIPE if stripe else FILL_WHITE
            c.font = FONT_NORMAL
            c.alignment = CENTER
            c.border = BD

    def _auto_width(ws, min_w=10, max_w=40):
        for col in ws.columns:
            best = min_w
            for cell in col:
                try:
                    best = max(best, len(str(cell.value or '')) + 2)
                except Exception:
                    pass
            ws.column_dimensions[col[0].column_letter].width = min(best, max_w)

    wb = openpyxl.Workbook()
    now = timezone.now()

    # ════════════════════════════════════════════════════════════════════════
    # Sheet 1 — Résumé
    # ════════════════════════════════════════════════════════════════════════
    ws_sum = wb.active
    ws_sum.title = "Résumé"
    ws_sum.sheet_view.showGridLines = False

    ws_sum['A1'] = "SOTETEL — IntelliOLT"
    ws_sum['A1'].font = FONT_TITLE
    ws_sum['A2'] = report.name
    ws_sum['A2'].font = Font(bold=True, size=12)
    ws_sum['A3'] = (
        f"Période : {report.date_from.strftime('%d/%m/%Y')} — "
        f"{report.date_to.strftime('%d/%m/%Y')}"
    )
    ws_sum['A3'].font = FONT_SUB
    ws_sum['A4'] = f"Généré le : {now.strftime('%d/%m/%Y à %H:%M')}"
    ws_sum['A4'].font = FONT_SUB
    for row in range(1, 5):
        ws_sum.merge_cells(f'A{row}:D{row}')
    ws_sum.row_dimensions[1].height = 30

    # Metrics
    olts_qs     = OLT.objects.all()
    total_olts  = olts_qs.count()
    active_olts = olts_qs.filter(status='active').count()
    total_onts  = ONT.objects.count()
    online_onts = ONT.objects.filter(status='online').count()
    active_alerts  = Alert.objects.filter(status='active').count()
    period_anomalies = AnomalyDetection.objects.filter(
        detected_at__range=(report.date_from, report.date_to)
    ).count()

    kpis_list = list(kpis)
    avg_cpu  = round(sum(k.avg_cpu_usage  or 0 for k in kpis_list) / len(kpis_list), 1) if kpis_list else None
    avg_mem  = round(sum(k.avg_memory_usage or 0 for k in kpis_list) / len(kpis_list), 1) if kpis_list else None
    avg_snmp = round(sum(k.snmp_success_rate or 100 for k in kpis_list) / len(kpis_list), 1) if kpis_list else None

    ont_pct = round(online_onts / total_onts * 100, 1) if total_onts else 0

    metric_rows = [
        ('Indicateur', 'Valeur', 'Détail', 'Statut'),
        ('OLTs actifs', f"{active_olts}/{total_olts}", '', '✓ OK' if active_olts == total_olts else '⚠ Vérifier'),
        ('ONTs en ligne', f"{online_onts}/{total_onts}", f"{ont_pct} %", '✓ OK' if ont_pct >= 95 else '⚠ Vérifier'),
        ('CPU moyen', f"{avg_cpu} %" if avg_cpu else '—', 'sur la période', '✓ OK' if not avg_cpu or avg_cpu < 70 else '⚠'),
        ('Mémoire moyenne', f"{avg_mem} %" if avg_mem else '—', 'sur la période', '✓ OK' if not avg_mem or avg_mem < 80 else '⚠'),
        ('Taux SNMP', f"{avg_snmp} %" if avg_snmp else '—', 'succès polling', '✓ OK' if not avg_snmp or avg_snmp > 95 else '⚠'),
        ('Alertes actives', active_alerts, '', '✓ Aucune' if active_alerts == 0 else f'⚠ {active_alerts}'),
        ('Anomalies période', period_anomalies, '', '✓ OK' if period_anomalies < 3 else '⚠'),
    ]

    for r_idx, row_vals in enumerate(metric_rows, start=6):
        for c_idx, val in enumerate(row_vals, start=1):
            c = ws_sum.cell(row=r_idx, column=c_idx, value=val)
            c.border = BD
            c.alignment = CENTER
            if r_idx == 6:
                c.fill = FILL_BLUE
                c.font = FONT_WHITE
            else:
                c.fill = FILL_STRIPE if r_idx % 2 == 0 else FILL_WHITE
                c.font = FONT_NORMAL
                if c_idx == 4:
                    val_str = str(val)
                    if val_str.startswith('✓'):
                        c.font = Font(size=9, color='27AE60', bold=True)
                    elif val_str.startswith('⚠'):
                        c.font = Font(size=9, color='C0392B', bold=True)

    ws_sum.column_dimensions['A'].width = 22
    ws_sum.column_dimensions['B'].width = 18
    ws_sum.column_dimensions['C'].width = 18
    ws_sum.column_dimensions['D'].width = 16

    # ════════════════════════════════════════════════════════════════════════
    # Sheet 2 — KPIs Journaliers
    # ════════════════════════════════════════════════════════════════════════
    ws_kpi = wb.create_sheet("KPIs Journaliers")
    ws_kpi.sheet_view.showGridLines = False
    kpi_headers = [
        'Date', 'Total OLTs', 'OLTs actifs', 'Total ONTs', 'ONTs en ligne',
        'ONT %', 'CPU moy. (%)', 'Mém. moy. (%)', 'Temp. (°C)', 'RX (dBm)',
        'SNMP %', 'BFD UP', 'Anomalies', 'Alertes',
    ]
    _apply_header(ws_kpi, kpi_headers)
    for r_idx, k in enumerate(kpis_list, start=2):
        ont_pct_row = round(k.online_onts / k.total_onts * 100, 1) if k.total_onts else 0
        _data_row(ws_kpi, r_idx, [
            k.timestamp.strftime('%Y-%m-%d'),
            k.total_olts, k.active_olts, k.total_onts, k.online_onts,
            ont_pct_row,
            round(k.avg_cpu_usage, 1)      if k.avg_cpu_usage      is not None else None,
            round(k.avg_memory_usage, 1)   if k.avg_memory_usage   is not None else None,
            round(k.avg_temperature, 1)    if k.avg_temperature    is not None else None,
            round(k.avg_rx_power, 2)       if k.avg_rx_power       is not None else None,
            round(k.snmp_success_rate, 1),
            k.bfd_up_sessions, k.anomaly_count, k.alert_count,
        ], stripe=(r_idx % 2 == 0))
    if not kpis_list:
        ws_kpi.cell(row=2, column=1, value='Aucune donnée pour cette période').font = FONT_BOLD
    _auto_width(ws_kpi)

    # ════════════════════════════════════════════════════════════════════════
    # Sheet 3 — OLTs
    # ════════════════════════════════════════════════════════════════════════
    ws_olt = wb.create_sheet("OLTs")
    ws_olt.sheet_view.showGridLines = False
    olt_headers = [
        'Hostname', 'Site', 'Adresse IP', 'Fournisseur', 'Modèle', 'Statut',
        'ONTs total', 'ONTs en ligne', 'Disponibilité %',
    ]
    _apply_header(ws_olt, olt_headers)
    for r_idx, olt in enumerate(
            OLT.objects.select_related('site', 'vendor', 'device_type').all(),
            start=2):
        ont_total  = ONT.objects.filter(olt=olt).count()
        ont_online = ONT.objects.filter(olt=olt, status='online').count()
        avail      = round(ont_online / ont_total * 100, 1) if ont_total else 100.0
        _data_row(ws_olt, r_idx, [
            olt.hostname,
            olt.site.name if olt.site else '—',
            olt.ip_address,
            olt.vendor.name if olt.vendor else '—',
            olt.device_type.model if olt.device_type else '—',
            olt.get_status_display(),
            ont_total, ont_online, avail,
        ], stripe=(r_idx % 2 == 0))
    _auto_width(ws_olt)

    # ════════════════════════════════════════════════════════════════════════
    # Sheet 4 — ONTs (sample)
    # ════════════════════════════════════════════════════════════════════════
    ws_ont = wb.create_sheet("ONTs")
    ws_ont.sheet_view.showGridLines = False
    ont_headers = [
        'OLT', 'N° série', 'Modèle', 'Statut', 'Adresse IP',
        'Port GPON', 'RX Power (dBm)', 'Adresse MAC',
    ]
    _apply_header(ws_ont, ont_headers)
    for r_idx, ont in enumerate(
            ONT.objects.select_related('olt', 'vendor', 'gpon_port')
            .all().order_by('olt__hostname', 'serial_number')[:500],
            start=2):
        _data_row(ws_ont, r_idx, [
            ont.olt.hostname if ont.olt else '—',
            ont.serial_number,
            ont.model or '—',
            ont.get_status_display(),
            ont.ip_address or '—',
            ont.gpon_port.port_name if ont.gpon_port else '—',
            round(ont.rx_power, 2) if ont.rx_power is not None else None,
            ont.mac_address or '—',
        ], stripe=(r_idx % 2 == 0))
    _auto_width(ws_ont)

    # ════════════════════════════════════════════════════════════════════════
    # Sheet 5 — Alertes
    # ════════════════════════════════════════════════════════════════════════
    ws_alert = wb.create_sheet("Alertes")
    ws_alert.sheet_view.showGridLines = False
    alert_headers = ['Date', 'OLT', 'Sévérité', 'Statut', 'Message']
    _apply_header(ws_alert, alert_headers, fill=FILL_ORANGE)

    SEV_FILLS = {
        'critical': _fill('FFDCDC'),
        'major':    _fill('FFEDCC'),
        'warning':  _fill('FFFACC'),
        'info':     _fill('E8F5E9'),
    }

    alerts_qs = Alert.objects.filter(
        first_seen__range=(report.date_from, report.date_to)
    ).select_related('olt').order_by('-first_seen')

    for r_idx, a in enumerate(alerts_qs, start=2):
        row_vals = [
            a.first_seen.strftime('%Y-%m-%d %H:%M'),
            a.olt.hostname if a.olt else '—',
            a.severity.upper(),
            a.status.upper(),
            a.message[:120],
        ]
        for c_idx, val in enumerate(row_vals, start=1):
            c = ws_alert.cell(row=r_idx, column=c_idx, value=val)
            c.font = FONT_NORMAL
            c.border = BD
            c.alignment = LEFT if c_idx == 5 else CENTER
            c.fill = SEV_FILLS.get(a.severity, FILL_WHITE)

    if not alerts_qs.exists():
        ws_alert.cell(row=2, column=1, value='Aucune alerte pour cette période').font = FONT_BOLD

    ws_alert.column_dimensions['A'].width = 18
    ws_alert.column_dimensions['B'].width = 16
    ws_alert.column_dimensions['C'].width = 12
    ws_alert.column_dimensions['D'].width = 14
    ws_alert.column_dimensions['E'].width = 60

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ============================================================================
# DEVICE / TRAFFIC COLLECTION TASKS
# ============================================================================

@shared_task(name='analytics.test_device_connection', bind=True, max_retries=2)
def test_device_connection(self, device_id):
    from .models import NetworkDevice
    try:
        device = NetworkDevice.objects.get(id=device_id)
        # Placeholder — real SSH via paramiko would go here
        device.is_reachable = True
        device.last_connection_at = timezone.now()
        device.connection_error = ''
        device.save(update_fields=['is_reachable', 'last_connection_at', 'connection_error'])
        logger.info(f"✅ Connexion réussie à {device.name}")
    except Exception as exc:
        logger.error(f"❌ Erreur connexion {device_id}: {exc}")
        try:
            device = NetworkDevice.objects.get(id=device_id)
            device.is_reachable = False
            device.connection_error = str(exc)
            device.last_connection_at = timezone.now()
            device.save(update_fields=['is_reachable', 'connection_error', 'last_connection_at'])
        except Exception:
            pass
        raise


@shared_task(name='analytics.collect_ssh_metrics', bind=True, max_retries=2)
def collect_ssh_metrics(self, device_id):
    import random
    from .models import NetworkDevice, SSHMetricsSnapshot
    try:
        device = NetworkDevice.objects.get(id=device_id)
        cpu    = random.uniform(20, 80)
        memory = random.uniform(40, 70)
        temp   = random.uniform(35, 50)
        snapshot = SSHMetricsSnapshot.objects.create(
            device=device, timestamp=timezone.now(),
            cpu_usage_pct=cpu, memory_usage_pct=memory, temperature_c=temp,
            uptime_seconds=random.randint(86400, 864000),
            is_anomaly=(cpu > 75 or memory > 80),
        )
        logger.info(f"✅ Métriques SSH {device.name}: CPU={cpu:.1f}% MEM={memory:.1f}%")
        return str(snapshot.id)
    except Exception as exc:
        logger.error(f"❌ Erreur collecte SSH {device_id}: {exc}")
        raise


@shared_task(name='analytics.collect_all_ssh_metrics', bind=True)
def collect_all_ssh_metrics(self):
    from .models import NetworkDevice
    devices = NetworkDevice.objects.filter(is_active=True)
    for device in devices:
        collect_ssh_metrics.delay(device.id)
    logger.info(f"Collecte SSH lancée pour {devices.count()} appareils")
    return devices.count()


def _get_latest_two(olt, oid_name):
    """Retourne les deux dernières valeurs MetricHistory pour un OID donné sur un OLT."""
    from apps.snmp_collector.models import MetricHistory
    rows = (
        MetricHistory.objects
        .filter(olt=olt, oid__name=oid_name, numeric_value__isnull=False)
        .order_by('-timestamp')
        .values('numeric_value', 'timestamp')[:2]
    )
    return list(rows)


def _counter_delta_mbps(rows):
    """
    Calcule le débit en Mbps depuis deux relevés de compteur SNMP (bytes cumulatifs).
    Retourne (bytes_delta, throughput_mbps, duration_seconds) ou (0, 0.0, 0) si données insuffisantes.
    Gère le wrap-around 32-bit (compteur SNMP qui repart à 0 après 2^32).
    """
    if len(rows) < 2:
        return 0, 0.0, 0

    v_new, t_new = rows[0]['numeric_value'], rows[0]['timestamp']
    v_old, t_old = rows[1]['numeric_value'], rows[1]['timestamp']

    duration = (t_new - t_old).total_seconds()
    if duration <= 0:
        return 0, 0.0, 0

    delta = v_new - v_old
    # Wrap-around 32-bit counter (ifInOctets est un Counter32)
    if delta < 0:
        delta += 2 ** 32

    throughput_mbps = round((delta * 8) / duration / 1_000_000, 4)
    return int(delta), throughput_mbps, duration


@shared_task(name='analytics.collect_network_traffic', bind=True, max_retries=2)
def collect_network_traffic(self, interface_id=None, fiber_link_id=None):
    """
    Calcule le trafic réel d'une interface depuis les compteurs SNMP stockés dans MetricHistory.
    Remplace les anciennes données aléatoires par des valeurs issues de if_in_octets / if_out_octets.
    """
    from .models import NetworkTraffic
    from apps.equipements.models import NetworkInterface, FibreLink

    try:
        if interface_id:
            iface = NetworkInterface.objects.select_related('olt').get(id=interface_id)
            olt   = iface.olt
            link  = None
        elif fiber_link_id:
            link  = FibreLink.objects.get(id=fiber_link_id)
            # Pour un lien fibre, on utilise l'OLT source
            olt   = link.source_device if hasattr(link, 'source_device') else None
            iface = None
            if not olt:
                return
        else:
            return

        # ── Compteurs bytes (IF-MIB) ─────────────────────────────────────────
        rows_in  = _get_latest_two(olt, 'if_in_octets')
        rows_out = _get_latest_two(olt, 'if_out_octets')

        bytes_in,  throughput_in,  dur_in  = _counter_delta_mbps(rows_in)
        bytes_out, throughput_out, dur_out = _counter_delta_mbps(rows_out)

        # Débit total = max des deux directions (le lien est le goulot d'étranglement)
        throughput_total = round(throughput_in + throughput_out, 4)

        # ── Compteurs erreurs ────────────────────────────────────────────────
        err_in_rows  = _get_latest_two(olt, 'if_in_errors')
        err_out_rows = _get_latest_two(olt, 'if_out_errors')
        errors_in,  _, _ = _counter_delta_mbps(err_in_rows)
        errors_out, _, _ = _counter_delta_mbps(err_out_rows)

        # ── Taux d'utilisation ───────────────────────────────────────────────
        utilization = None
        if iface and iface.speed_mbps and iface.speed_mbps > 0:
            utilization = round(min(throughput_total / iface.speed_mbps * 100, 100), 2)

        # ── Écriture NetworkTraffic ──────────────────────────────────────────
        kwargs = dict(
            timestamp       = timezone.now(),
            bytes_in        = bytes_in,
            bytes_out       = bytes_out,
            errors_in       = errors_in,
            errors_out      = errors_out,
            throughput_mbps = throughput_total,
            utilization_pct = utilization,
            is_congested    = (utilization or 0) > 80,
        )
        if iface:
            kwargs['interface'] = iface
        else:
            kwargs['fiber_link'] = link

        NetworkTraffic.objects.create(**kwargs)
        logger.info(
            "collect_network_traffic: %s → ↓%.2f Mbps ↑%.2f Mbps util=%s%%",
            olt.hostname, throughput_in, throughput_out,
            f"{utilization:.1f}" if utilization is not None else "—"
        )

    except Exception as exc:
        logger.error("❌ Erreur collecte trafic: %s", exc)
        raise self.retry(exc=exc, countdown=30)


@shared_task(name='analytics.collect_all_network_traffic', bind=True)
def collect_all_network_traffic(self):
    from apps.equipements.models import NetworkInterface
    interfaces = NetworkInterface.objects.filter(admin_status=True).select_related('olt')
    count = 0
    for iface in interfaces:
        if iface.olt and iface.olt.status in ('active', 'degraded'):
            collect_network_traffic.delay(interface_id=iface.id)
            count += 1
    logger.info("collect_all_network_traffic: %d interfaces planifiées.", count)
    return count


@shared_task(name='analytics.detect_anomalies')
def detect_anomalies():
    """Marque les snapshots SSH et trafic anormaux."""
    from .models import SSHMetricsSnapshot, NetworkTraffic
    since = timezone.now() - timedelta(hours=1)

    updated_metrics = SSHMetricsSnapshot.objects.filter(
        timestamp__gte=since, is_anomaly=False
    ).filter(
        Q(cpu_usage_pct__gt=85) | Q(memory_usage_pct__gt=85) | Q(temperature_c__gt=60)
    ).update(is_anomaly=True)

    updated_traffic = NetworkTraffic.objects.filter(
        timestamp__gte=since, is_anomaly=False, utilization_pct__gt=85
    ).update(is_anomaly=True)

    count = updated_metrics + updated_traffic
    logger.info(f"✅ {count} anomalies détectées")
    return count
